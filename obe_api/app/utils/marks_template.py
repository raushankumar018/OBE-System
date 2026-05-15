"""
marks_template.py
─────────────────
Utility endpoint to download a pre-filled Excel template
for entering student marks — uses actual question bank IDs from the session.
"""

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
import os, uuid
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.config import get_settings
from app.database import co_sessions_col
from app.services.obe_engine import build_question_bank

router   = APIRouter(prefix="/template", tags=["Templates"])
settings = get_settings()


@router.get("/marks/{session_id}", summary="Download blank marks entry Excel template")
async def download_marks_template(session_id: str):
    """
    Download a blank Excel template pre-filled with:
    - All Question_IDs for this session's exam structure
    - Student roll number rows (one row per student per question)
    - Instructions sheet

    Fill in the Marks_Obtained column and upload to POST /attainment/calculate.
    """
    session = await co_sessions_col().find_one({"session_id": session_id})
    if not session:
        raise HTTPException(404, f"Session {session_id} not found.")

    cfg = {
        "num_modules": session.get("num_modules", 2),
        "t1_parta_bits": session.get("t1_parta_bits", {"a": 2, "b": 4, "c": 4}),
        "t1_partb_bits": session.get("t1_partb_bits", {"a": 2, "b": 4, "c": 4}),
        "t2_max": session.get("t2_max", 5),
        "t3_max": session.get("t3_max", 5),
        "t4_mcq_count": session.get("t4_mcq_count", 10),
        "t4_mcq_marks": session.get("t4_mcq_marks", 0.5),
        "t4_desc_count": session.get("t4_desc_count", 3),
        "t4_desc_marks": session.get("t4_desc_marks", 5),
        "t4_desc_bits": session.get("t4_desc_bits", {"a": 2, "b": 2, "c": 1}),
        "t5_cla_count": session.get("t5_cla_count", 4),
        "t5_cla_marks": session.get("t5_cla_marks", 20),
        "lab_ext_max": session.get("lab_ext_max", 20),
        "ete_parta_count": session.get("ete_parta_count", 4),
        "ete_parta_marks": session.get("ete_parta_marks", 8),
        "ete_parta_bits": session.get("ete_parta_bits", {"a": 3, "b": 3, "c": 2}),
        "ete_partb_count": session.get("ete_partb_count", 2),
        "ete_partb_marks": session.get("ete_partb_marks", 14),
        "ete_partb_bits": session.get("ete_partb_bits", {"a": 5, "b": 5, "c": 4}),
    }

    q_bank      = build_question_bank(cfg)
    num_students = session.get("num_students", 60)
    course_code  = session["course_code"]
    students     = [f"{course_code[:4]}{str(i+1).zfill(3)}" for i in range(num_students)]

    # Build template rows
    rows = []
    for roll in students:
        for _, q in q_bank.iterrows():
            rows.append({
                "Roll_No":        roll,
                "Question_ID":    q["Question_ID"],
                "Exam":           q["Exam"],
                "Module":         q["Module"],
                "Component":      q["Component"],
                "Max_Marks":      q["Max_Marks"],
                "Marks_Obtained": "",    # ← Faculty fills this in
                "Bloom_Level":    q["Bloom_Level"],
                "Mapped_CO":      q.get("Mapped_CO", ""),
            })

    df = pd.DataFrame(rows)

    # Write Excel
    os.makedirs(settings.output_dir, exist_ok=True)
    out_path = os.path.join(settings.output_dir, f"marks_template_{session_id[:8]}.xlsx")
    wb = openpyxl.Workbook()

    # ── Sheet 1: Instructions ──────────────────────────────────────────────
    ws_inst = wb.active; ws_inst.title = "Instructions"
    instructions = [
        ["OBE Marks Entry Template"],
        [f"Course: {session['course_code']} — {session['course_name']}"],
        [f"Academic Year: {session['academic_year']}"],
        [""],
        ["HOW TO USE:"],
        ["1. Go to the 'Marks' sheet"],
        ["2. Fill in the 'Marks_Obtained' column for each student and question"],
        ["3. Do NOT modify Roll_No, Question_ID, or Max_Marks columns"],
        ["4. For Module Bank questions: only fill rows for the question assigned to each student"],
        ["   (each student has a different module bank question — MB1 to MB10)"],
        ["5. Leave other module bank rows empty — system will ignore them"],
        ["6. Save and upload via POST /attainment/calculate"],
        [""],
        ["QUESTION ID FORMAT:"],
        ["  M1_MB3_a    = Module 1, Module Bank Q3, Bit A"],
        ["  M1_T1B_b    = Module 1, T1 Part B, Bit B"],
        ["  M1_Viva     = Module 1, T2 Viva"],
        ["  M1_DocPPT   = Module 1, T3 Document+PPT"],
        ["  M1_MCQ5     = Module 1, T4 MCQ Question 5"],
        ["  M1_T4Q2_c   = Module 1, T4 Descriptive Q2 Bit C"],
        ["  M1_CLA3     = Module 1, T5 CLA Assignment 3"],
        ["  Lab_External = Lab External Exam"],
        ["  ETE_A2_b    = End Term Part A, Q2 Bit B"],
        ["  ETE_B1_a    = End Term Part B, Q1 Bit A"],
    ]
    for i, row in enumerate(instructions, 1):
        ws_inst.cell(i, 1, row[0] if row else "")
        if i == 1:
            ws_inst.cell(i, 1).font = Font(bold=True, size=13, color="1F3864")
        elif row and row[0].startswith("HOW") or (row and row[0].startswith("QUESTION")):
            ws_inst.cell(i, 1).font = Font(bold=True, size=11)
    ws_inst.column_dimensions["A"].width = 80

    # ── Sheet 2: Marks Entry ───────────────────────────────────────────────
    ws_marks = wb.create_sheet("Marks")
    headers = list(df.columns)
    header_fill = PatternFill("solid", fgColor="1F3864")
    for ci, h in enumerate(headers, 1):
        cell = ws_marks.cell(1, ci, h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fill-in column highlight
    fill_col_idx = headers.index("Marks_Obtained") + 1
    fill_highlight = PatternFill("solid", fgColor="FFFACD")

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws_marks.cell(ri, ci, val)
            cell.alignment = Alignment(horizontal="center")
            if ci == fill_col_idx:
                cell.fill = fill_highlight

    # Auto-width
    for col in ws_marks.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws_marks.column_dimensions[
            openpyxl.utils.get_column_letter(col[0].column)
        ].width = min(max(max_len + 2, 10), 35)

    # Freeze header row
    ws_marks.freeze_panes = "A2"

    wb.save(out_path)
    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"marks_template_{course_code}_{session['academic_year']}.xlsx"
    )
