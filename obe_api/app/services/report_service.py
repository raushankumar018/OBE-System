"""
report_service.py
─────────────────
Generates PDF and Excel reports from stored attainment data.
Used both after live calculation AND for re-downloading from MongoDB.
"""

import os
import io
from typing import Dict, List, Optional
from datetime import datetime

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
import numpy as np

from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLOOM_COLORS = {
    "Remember": "#B5D4F4", "Understand": "#A9DFBF", "Apply": "#F9E79F",
    "Analyze":  "#F0B27A", "Evaluate":   "#F1948A", "Create": "#C39BD3",
}
STRENGTH_COLORS = {"High": "FFD700", "Moderate": "FFFACD", "Low": "E8F4E8", "--": "FFFFFF"}


# ════════════════════════════════════════════════════════════════════════════
# DASHBOARD CHART
# ════════════════════════════════════════════════════════════════════════════

def generate_dashboard_image(attainment_data: Dict, output_path: str) -> str:
    co_att  = pd.DataFrame(attainment_data["co_attainment"])
    po_att  = pd.DataFrame(attainment_data["po_attainment"])
    stu_piv = pd.DataFrame(attainment_data["student_pivot"])
    co_student = pd.DataFrame(attainment_data["co_student"])
    meta    = attainment_data.get("metadata", {})

    co_cols    = [c["co_id"] for c in attainment_data["cos"]]
    target     = meta.get("target_attainment", 1.67)
    course_str = f"{meta.get('course_code','')} — {meta.get('course_name','')}"

    fig = plt.figure(figsize=(20, 12))
    fig.suptitle(f"CO-PO-PSO Attainment Dashboard\n{course_str} | {meta.get('university','')} | {meta.get('academic_year','')}",
                 fontsize=13, fontweight="bold")
    axes = [fig.add_subplot(2, 3, i) for i in range(1, 7)]

    # Panel 1: CO Final Score vs Target
    ax = axes[0]
    co_ids = co_att["CO_ID"].tolist()
    scores = co_att["Final_Score"].tolist()
    colors = ["#2ECC71" if s >= target else "#E74C3C" for s in scores]
    bars = ax.bar(co_ids, scores, color=colors, width=0.55, edgecolor="white")
    ax.axhline(target, color="black", ls="--", lw=1.8, label=f"Target={target}")
    ax.set_ylim(0, 3.4); ax.set_ylabel("Score (0–3)")
    ax.set_title("CO Attainment Score", fontweight="bold"); ax.legend(fontsize=8)
    for bar, s in zip(bars, scores):
        col = "#1A5276" if s >= target else "#C0392B"
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.06,
                f"{s:.2f}", ha="center", fontsize=9, fontweight="bold", color=col)

    # Panel 2: FA vs SA
    ax = axes[1]
    x = np.arange(len(co_ids))
    ax.plot(x, co_att["IA_Pct"],   "o-", color="#3498DB", lw=2, ms=7, label="FA%")
    ax.plot(x, co_att["ETE_Pct"],  "s-", color="#E67E22", lw=2, ms=7, label="SA%")
    ax.plot(x, co_att["Mean_Pct"], "^--", color="#9B59B6", lw=1.5, ms=5, label="Overall%")
    ax.axhline(60, color="gray", ls=":", lw=1.2)
    ax.set_xticks(x); ax.set_xticklabels(co_ids); ax.set_ylim(0, 105)
    ax.set_title("FA vs SA Performance", fontweight="bold"); ax.legend(fontsize=8)

    # Panel 3: PO Attainment bar
    ax = axes[2]
    if not po_att.empty:
        po_pcts   = po_att["attainment_pct"].tolist()
        po_labels = po_att["outcome"].tolist()
        po_colors = ["#2ECC71" if v >= 60 else "#E74C3C" for v in po_pcts]
        bars3 = ax.bar(po_labels, po_pcts, color=po_colors, edgecolor="white")
        ax.axhline(60, color="red", ls="--", lw=1.2); ax.set_ylim(0, 110)
        ax.set_title("PO Attainment %", fontweight="bold"); ax.set_ylabel("Attainment %")
        ax.tick_params(axis="x", rotation=45)
        for b, v in zip(bars3, po_pcts):
            ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 1.5,
                    f"{v:.0f}%", ha="center", fontsize=8, fontweight="bold")

    # Panel 4: Student level stacked bar
    ax = axes[3]
    lvl_colors = {"Level 1": "#E74C3C", "Level 2": "#F39C12", "Level 3": "#2ECC71"}
    bottom = np.zeros(len(co_ids))
    for key, col in lvl_colors.items():
        n = int(key[-1])
        vals = co_att[f"Students_Level{n}"].tolist()
        ax.bar(co_ids, vals, bottom=bottom, color=col, label=key, edgecolor="white")
        bottom += np.array(vals)
    ax.set_ylabel("Students"); ax.set_title("Student Level Distribution", fontweight="bold")
    ax.legend(fontsize=8)

    # Panel 5: Overall grade pie
    ax = axes[4]
    grade_counts = pd.Series([r["Grade"] for r in attainment_data["student_pivot"]]).value_counts()
    grade_colors = {"O": "#1A5276", "A+": "#2980B9", "A": "#5DADE2",
                    "B+": "#2ECC71", "B": "#F39C12", "F": "#E74C3C"}
    pie_colors = [grade_colors.get(g, "#95A5A6") for g in grade_counts.index]
    ax.pie(grade_counts.values, labels=grade_counts.index, colors=pie_colors,
           autopct="%1.0f%%", startangle=90, wedgeprops={"edgecolor": "white"})
    ax.set_title("Grade Distribution", fontweight="bold")

    # Panel 6: Bloom distribution in COs
    ax = axes[5]
    bloom_order = ["Remember", "Understand", "Apply", "Analyze", "Evaluate", "Create"]
    bloom_map   = {c["co_id"]: c["bloom_level"] for c in attainment_data["cos"]}
    b_counts = {l: sum(1 for v in bloom_map.values() if v == l) for l in bloom_order}
    b_in_order = [l for l in bloom_order if b_counts[l] > 0]
    bars6 = ax.barh(b_in_order, [b_counts[l] for l in b_in_order],
                    color=[BLOOM_COLORS.get(l, "#ccc") for l in b_in_order], edgecolor="white")
    ax.set_title("Bloom Level Distribution", fontweight="bold")
    for b, l in zip(bars6, b_in_order):
        if b_counts[l] > 0:
            ax.text(b.get_width() + 0.05, b.get_y() + b.get_height() / 2,
                    str(b_counts[l]), va="center", fontsize=10, fontweight="bold")

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()
    return output_path


# ════════════════════════════════════════════════════════════════════════════
# PDF REPORT
# ════════════════════════════════════════════════════════════════════════════

def sf(v, n=80):
    return str(v).encode("latin-1", "replace").decode("latin-1")[:n]


class OBEReport(FPDF):
    def __init__(self, meta: Dict):
        super().__init__(orientation="L", unit="mm", format="A4")
        self.meta = meta

    def header(self):
        self.set_font("Helvetica", "B", 9)
        self.set_fill_color(31, 56, 100); self.set_text_color(255, 255, 255)
        title = sf(f"OBE Report | {self.meta.get('course_code','')} — {self.meta.get('course_name','')} | {self.meta.get('university','')} | {self.meta.get('academic_year','')}")
        self.cell(0, 7, title, fill=True, ln=1, align="C")
        self.set_text_color(0); self.ln(1)

    def footer(self):
        self.set_y(-10)
        self.set_font("Helvetica", "I", 7); self.set_text_color(150)
        self.cell(0, 7, f"Page {self.page_no()} | AI CO-PO-PSO Attainment System | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", align="C")

    def section(self, title: str):
        self.set_font("Helvetica", "B", 10)
        self.set_fill_color(31, 56, 100); self.set_text_color(255, 255, 255)
        self.cell(0, 7, sf(title), fill=True, ln=1)
        self.set_text_color(0); self.ln(1)

    def tbl(self, headers, rows, widths, rh=5):
        self.set_font("Helvetica", "B", 8); self.set_fill_color(220, 220, 230)
        for h, w in zip(headers, widths):
            self.cell(w, 6, sf(h), border=1, fill=True, align="C")
        self.ln()
        self.set_font("Helvetica", "", 7)
        for row in rows:
            if self.get_y() > 182:
                self.add_page()
            for v, w in zip(row, widths):
                self.cell(w, rh, sf(v), border=1)
            self.ln()
        self.ln(2)


def generate_pdf_report(attainment_data: Dict, output_path: str,
                         dashboard_img_path: Optional[str] = None) -> str:
    meta = attainment_data.get("metadata", {})
    pdf  = OBEReport(meta)
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    # ── Exam Structure ────────────────────────────────────────────────────
    pdf.section("1. Course Information")
    info_rows = [
        ["University", meta.get("university", "")],
        ["Department", meta.get("department", "")],
        ["Course", f"{meta.get('course_code','')} — {meta.get('course_name','')}"],
        ["Semester", meta.get("semester", "")],
        ["Academic Year", meta.get("academic_year", "")],
        ["Students", str(meta.get("num_students", ""))],
        ["Target Attainment", f"{meta.get('target_attainment', 1.67)}/3"],
    ]
    pdf.tbl(["Field", "Value"], info_rows, [60, 210])

    # ── Course Outcomes ────────────────────────────────────────────────────
    pdf.section("2. Course Outcomes")
    cos = attainment_data.get("cos", [])
    pdf.tbl(
        ["CO", "Unit", "Bloom", "CO Statement", "PO Mapping"],
        [[c["co_id"], str(c["unit_no"]), c["bloom_level"],
          c["co_text"][:90],
          ", ".join(c["po_mapping"]) if isinstance(c["po_mapping"], list) else str(c["po_mapping"])]
         for c in cos],
        [14, 14, 20, 178, 44]
    )

    # ── CO Attainment ──────────────────────────────────────────────────────
    pdf.section("3. CO Attainment (Formative Assessment (FA) + Summative Assessment (SA))")
    co_att = attainment_data["co_attainment"]
    pdf.tbl(
        ["CO", "FA%", "SA%", "DA", "IDA", "Final", "Level", "Status"],
        [[r["CO_ID"], f"{r['IA_Pct']:.1f}%", f"{r['ETE_Pct']:.1f}%",
          f"{r['DA_Score']:.2f}", f"{r['IDA_Score']:.2f}", f"{r['Final_Score']:.2f}",
          f"Lv{r['Class_Level']}", r["Attained"]]
         for r in co_att],
        [18, 22, 22, 22, 22, 22, 18, 34]
    )

    # ── PO Attainment ──────────────────────────────────────────────────────
    pdf.section("4. Program Outcome Attainment")
    po_att = attainment_data["po_attainment"]
    if po_att:
        pdf.tbl(
            ["PO/PSO", "Score", "Level", "Attainment%", "Contributors"],
            [[r["outcome"], f"{r['score']:.2f}", f"Lv{r['level']}",
              f"{r['attainment_pct']:.1f}%", r["contributors"][:40]]
             for r in po_att],
            [20, 22, 18, 26, 184]
        )

    # ── Student-wise CO Attainment ─────────────────────────────────────────
    pdf.section("5. Student-wise CO Attainment")
    co_cols = [c["co_id"] for c in cos]
    stu_piv = attainment_data["student_pivot"]
    pdf.tbl(
        ["S.No", "Roll No"] + co_cols + ["Overall%", "Grade"],
        [[str(i + 1), r["Roll_No"]] +
         [f"{r.get(cid, 0) or 0:.1f}" for cid in co_cols] +
         [f"{r.get('Overall_Pct', 0):.1f}%", r.get("Grade", "")]
         for i, r in enumerate(stu_piv)],
        [10, 22] + [20] * len(co_cols) + [24, 14]
    )

    # ── Dashboard Image ────────────────────────────────────────────────────
    if dashboard_img_path and os.path.exists(dashboard_img_path):
        pdf.add_page()
        pdf.section("6. Attainment Dashboard")
        pdf.image(dashboard_img_path, x=5, y=pdf.get_y() + 2, w=280)

    pdf.output(output_path)
    return output_path


# ════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ════════════════════════════════════════════════════════════════════════════

def _bdr():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _cs(c, bold=False, bg=None, fg="000000", align="center", size=10, wrap=False):
    c.font      = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = _bdr()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)


def _aw(ws, min_w=8, max_w=50):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)


def _excel_value(value):
    if value is None:
        return ""
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float) and pd.isna(value):
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return ", ".join(f"{k}: {v}" for k, v in value.items())
    return value


def generate_excel_report(attainment_data: Dict, output_path: str) -> str:
    meta   = attainment_data.get("metadata", {})
    cos    = attainment_data.get("cos", [])
    co_att = pd.DataFrame(attainment_data.get("co_attainment", []))
    po_att = pd.DataFrame(attainment_data.get("po_attainment", []))
    pso_att = pd.DataFrame(attainment_data.get("pso_attainment", []))
    stu_piv = pd.DataFrame(attainment_data.get("student_pivot", []))
    co_cols = [c["co_id"] for c in cos]
    matrix_label = attainment_data.get("matrix_label", {})
    matrix_num   = attainment_data.get("matrix_num", {})
    all_pos  = [f"PO{i}" for i in range(1, 13)]
    all_psos = list({r["outcome"] for r in attainment_data.get("pso_attainment", [])})
    all_cols = all_pos + all_psos

    wb = openpyxl.Workbook()

    # ── Sheet 1: Course Info ───────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Course Info"
    info = [("UNIVERSITY", meta.get("university", "")), ("DEPARTMENT", meta.get("department", "")),
            ("DEGREE", meta.get("degree", "")), ("COURSE CODE", meta.get("course_code", "")),
            ("COURSE NAME", meta.get("course_name", "")), ("SEMESTER", meta.get("semester", "")),
            ("ACADEMIC YEAR", meta.get("academic_year", "")),
            ("COORDINATOR", meta.get("coordinator", "")),
            ("NUM STUDENTS", meta.get("num_students", "")),
            ("TARGET ATTAINMENT", f"{meta.get('target_attainment', 1.67)}/3")]
    for i, (k, v) in enumerate(info, 1):
        _cs(ws1.cell(i, 1, k), bold=True, bg="1F3864", fg="FFFFFF")
        _cs(ws1.cell(i, 2, _excel_value(v)), align="left")
    _aw(ws1)

    # ── Sheet 2: Course Outcomes ───────────────────────────────────────────
    ws2 = wb.create_sheet("Course Outcomes")
    for ci, h in enumerate(["CO", "Unit", "Bloom", "CO Statement", "PO Mapping", "PSO Mapping"], 1):
        _cs(ws2.cell(1, ci, h), bold=True, bg="1F3864", fg="FFFFFF")
    for ri, c in enumerate(cos, 2):
        _cs(ws2.cell(ri, 1, _excel_value(c.get("co_id", ""))), bold=True)
        unit_no = c.get("unit_no", "")
        _cs(ws2.cell(ri, 2, _excel_value(f"Unit {unit_no}" if unit_no != "" else "")))
        _cs(ws2.cell(ri, 3, _excel_value(c.get("bloom_level", ""))), bg="E2EFDA")
        ws2.cell(ri, 4, _excel_value(c.get("co_text", ""))); _cs(ws2.cell(ri, 4), align="left", wrap=True)
        po_mapping = c.get("po_mapping", [])
        pso_mapping = c.get("pso_mapping", [])
        _cs(ws2.cell(ri, 5, _excel_value(po_mapping)))
        _cs(ws2.cell(ri, 6, _excel_value(pso_mapping)))
    ws2.column_dimensions["D"].width = 65; _aw(ws2)

    # ── Sheet 3: CO-PO-PSO Mapping Matrix ─────────────────────────────────
    ws3 = wb.create_sheet("CO-PO-PSO Mapping")
    _cs(ws3.cell(1, 1, "CO"), bold=True, bg="1F3864", fg="FFFFFF")
    for ci, col in enumerate(all_cols, 2):
        _cs(ws3.cell(1, ci, col), bold=True,
            bg="1F3864" if "PO" in col else "C00000", fg="FFFFFF")
    SCOL = {"High": "FFD700", "Moderate": "FFFACD", "Low": "E8F4E8", "--": "FFFFFF"}
    for ri, co in enumerate(cos, 2):
        _cs(ws3.cell(ri, 1, _excel_value(co.get("co_id", ""))), bold=True, bg="1F3864", fg="FFFFFF")
        for ci, col in enumerate(all_cols, 2):
            val = matrix_label.get(co["co_id"], {}).get(col, "--")
            _cs(ws3.cell(ri, ci, _excel_value(val)), bg=SCOL.get(val, "FFFFFF"))
    _aw(ws3)

    # ── Sheet 4: CO Attainment ─────────────────────────────────────────────
    ws4 = wb.create_sheet("CO Attainment")
    h4  = ["CO", "Bloom", "FA%", "SA%", "DA Score", "IDA Score", "Final Score", "Level", "Status", "CO Statement"]
    for ci, h in enumerate(h4, 1):
        _cs(ws4.cell(1, ci, h), bold=True, bg="1F3864", fg="FFFFFF")
    for ri, r in enumerate(co_att.to_dict("records"), 2):
        vals = [r["CO_ID"], r.get("Bloom_Level",""), round(r["IA_Pct"],1), round(r["ETE_Pct"],1),
                round(r["DA_Score"],2), round(r["IDA_Score"],2), round(r["Final_Score"],2),
                f"Level {r['Class_Level']}", r["Attained"], r.get("CO_Text","")[:80]]
        for ci, v in enumerate(vals, 1):
            bg = None
            if ci == 9:
                bg = "C6EFCE" if v == "Attained" else "FFC7CE"
            _cs(ws4.cell(ri, ci, _excel_value(v)), bg=bg, bold=(ci == 9),
                align="left" if ci == 10 else "center")
    _aw(ws4)

    # ── Sheet 5: Student Scores ────────────────────────────────────────────
    ws5 = wb.create_sheet("Student Scores")
    hdr5 = ["S.No", "Roll No"] + co_cols + ["Overall%", "Grade", "Level"]
    for ci, h in enumerate(hdr5, 1):
        _cs(ws5.cell(1, ci, h), bold=True, bg="1F3864", fg="FFFFFF")
    for ri, row in enumerate(stu_piv.to_dict("records"), 2):
        _cs(ws5.cell(ri, 1, ri - 1))
        _cs(ws5.cell(ri, 2, _excel_value(row.get("Roll_No", ""))), align="left")
        for ci, cid in enumerate(co_cols, 3):
            val = row.get(cid, 0) or 0
            bg  = "C6EFCE" if val >= 60 else ("FFEB9C" if val >= 40 else "FFC7CE")
            _cs(ws5.cell(ri, ci, _excel_value(round(val, 2))), bg=bg)
        ov  = row.get("Overall_Pct", 0) or 0
        tc  = 3 + len(co_cols)
        _cs(ws5.cell(ri, tc, _excel_value(round(ov, 2))))
        _cs(ws5.cell(ri, tc + 1, _excel_value(row.get("Grade", ""))))
        _cs(ws5.cell(ri, tc + 2, _excel_value(f"Level {_att_level(ov)}")))
    # Summary rows
    lr = 1 + len(stu_piv)
    _cs(ws5.cell(lr + 2, 1, "CLASS AVG%"), bold=True)
    for ci, cid in enumerate(co_cols, 3):
        avg = stu_piv[cid].mean() if cid in stu_piv else 0
        _cs(ws5.cell(lr + 2, ci, _excel_value(round(avg, 2))), bold=True, bg="BDD7EE")
    _aw(ws5)

    # ── Sheet 6: PO-PSO Attainment ─────────────────────────────────────────
    ws6 = wb.create_sheet("PO-PSO Attainment")
    for ci, h in enumerate(["Outcome", "Score (0-3)", "Attainment%", "Level", "Contributors", "Description"], 1):
        _cs(ws6.cell(1, ci, h), bold=True, bg="1F3864", fg="FFFFFF")
    ri = 2
    _cs(ws6.cell(ri, 1, "PROGRAM OUTCOMES"), bold=True, bg="BDD7EE"); ri += 1
    for r in po_att.to_dict("records"):
        for ci, val in enumerate([r["outcome"], r["score"], f"{r['attainment_pct']}%",
                                   f"Level {r['level']}", r["contributors"], r["description"]], 1):
            _cs(ws6.cell(ri, ci, _excel_value(val)), align="left" if ci in [5, 6] else "center")
        ri += 1
    if not pso_att.empty:
        _cs(ws6.cell(ri, 1, "PROGRAM SPECIFIC OUTCOMES"), bold=True, bg="E2EFDA"); ri += 1
        for r in pso_att.to_dict("records"):
            for ci, val in enumerate([r["outcome"], r["score"], f"{r['attainment_pct']}%",
                                       f"Level {r['level']}", r["contributors"], r["description"]], 1):
                _cs(ws6.cell(ri, ci, _excel_value(val)), align="left" if ci in [5, 6] else "center")
            ri += 1
    _aw(ws6)

    wb.save(output_path)
    return output_path


def _att_level(pct: float) -> int:
    thresholds = {1: (0, 40), 2: (40, 60), 3: (60, 101)}
    for lvl, (lo, hi) in thresholds.items():
        if lo <= pct < hi:
            return lvl
    return 3
