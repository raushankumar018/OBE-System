import os
import re
from typing import Dict, List, Optional

import openpyxl
from openpyxl import load_workbook


ROLL_RE = re.compile(r"^[0-9A-Z]{8,}$")


def attainment_level(pct: float) -> int:
    if pct < 40:
        return 1
    if pct < 60:
        return 2
    return 3


def is_vfstr_marks_workbook(path: str) -> bool:
    try:
        wb = load_workbook(path, data_only=False)
        ws = wb[wb.sheetnames[0]]
        return (
            str(ws["A1"].value or "").strip().upper() == "VFSTR"
            and str(ws["A6"].value or "").strip().upper() == "REGD.NO."
        )
    except Exception:
        return False


def _num(value) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def _safe_text(value) -> str:
    return str(value or "").strip()


def _sheet_rows(ws) -> List[Dict]:
    rows = []
    row_no = 9
    while row_no <= ws.max_row:
        reg_no = _safe_text(ws[f"A{row_no}"].value)
        if ROLL_RE.match(reg_no):
            rows.append(
                {
                    "reg_no": reg_no,
                    "m1_t1_a": _num(ws[f"B{row_no}"].value),
                    "m1_t1_b": _num(ws[f"C{row_no}"].value),
                    "m1_t4": _num(ws[f"AD{row_no}"].value),
                    "m1_qno_a": _num(ws[f"M{row_no}"].value),
                    "m1_qno_b": _num(ws[f"U{row_no}"].value),
                    "m2_t1_a": _num(ws[f"AE{row_no}"].value),
                    "m2_t1_b": _num(ws[f"AF{row_no}"].value),
                    "m2_t4": _num(ws[f"BE{row_no}"].value),
                    "m2_qno_a": _num(ws[f"AN{row_no}"].value),
                    "m2_qno_b": _num(ws[f"AV{row_no}"].value),
                    "sa_q1": _num(ws[f"BI{row_no}"].value),
                    "sa_q2": _num(ws[f"BJ{row_no}"].value),
                    "sa_q3": _num(ws[f"BK{row_no}"].value),
                    "sa_q4": _num(ws[f"BL{row_no}"].value),
                    "sa_q5": _num(ws[f"BM{row_no}"].value),
                    "sa_q6": _num(ws[f"BN{row_no}"].value),
                    "sa_practical": _num(ws[f"BO{row_no}"].value),
                }
            )
        row_no += 1
    return rows


def parse_vfstr_marks_workbook(path: str) -> Dict:
    wb = load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    course_line = _safe_text(ws["A3"].value)
    course_match = re.search(r"^(.*)\(([^()]+)\)$", course_line)
    if course_match:
        course_name = course_match.group(1).strip()
        course_code = course_match.group(2).strip()
    else:
        course_name = course_line
        course_code = ""

    header_line = _safe_text(ws["A2"].value)
    section_match = re.search(r"SECTION\s*-\s*([A-Za-z0-9]+)", header_line, re.IGNORECASE)
    faculty_line = _safe_text(ws["A4"].value)
    faculty_name = faculty_line.split(":", 1)[-1].strip() if ":" in faculty_line else faculty_line

    return {
        "course_code": course_code,
        "course_name": course_name,
        "header_line": header_line,
        "faculty_name": faculty_name,
        "section": section_match.group(1).strip() if section_match else "1",
        "rows": _sheet_rows(ws),
    }


def _tool_to_scale(score: Optional[float], max_marks: float) -> Optional[float]:
    if score is None:
        return None
    return max(0.0, min(3.0, (score / max_marks) * 3.0))


def _mean(values: List[Optional[float]]) -> float:
    nums = [v for v in values if v is not None]
    if not nums:
        return 0.0
    return sum(nums) / len(nums)


def build_vfstr_attainment_data(
    parsed: Dict,
    session: Dict,
    mapping_doc: Dict,
    indirect_score: float,
) -> Dict:
    cos = session["cos"]
    co_ids = [co["co_id"] for co in cos]
    co_text_map = {co["co_id"]: co.get("co_text", "") for co in cos}
    bloom_map = {co["co_id"]: co.get("bloom_level", "") for co in cos}
    matrix_num = mapping_doc["matrix_num"]
    po_dict = session.get("po_dict", {})
    pso_dict = session.get("pso_dict", {})
    target = float(session.get("target_attainment", 1.67))

    survey_score = indirect_score
    if survey_score > 3:
        survey_score = (survey_score / 5.0) * 3.0
    survey_score = max(0.0, min(3.0, survey_score))

    student_pivot = []
    co_buckets = {
        co_id: {"formative": [], "summative": [], "final": []}
        for co_id in co_ids
    }

    theory_max = {"CO1": 10.0, "CO2": 10.0, "CO3": 10.0, "CO4": 10.0, "CO5": 20.0, "CO6": 20.0}

    for row in parsed["rows"]:
        m1_a = _tool_to_scale(row["m1_t1_a"], 10.0)
        m1_b = _tool_to_scale(row["m1_t1_b"], 10.0)
        m1_t4 = _tool_to_scale((row["m1_t4"] or 0.0) * 2.0 if row["m1_t4"] is not None else None, 20.0)
        m2_a = _tool_to_scale(row["m2_t1_a"], 10.0)
        m2_b = _tool_to_scale(row["m2_t1_b"], 10.0)
        m2_t4 = _tool_to_scale((row["m2_t4"] or 0.0) * 2.0 if row["m2_t4"] is not None else None, 20.0)
        practical = _tool_to_scale(row["sa_practical"], 40.0)

        co_vals = {"Roll_No": row["reg_no"]}
        theory_scores = {
            "CO1": row["sa_q1"],
            "CO2": row["sa_q2"],
            "CO3": row["sa_q3"],
            "CO4": row["sa_q4"],
            "CO5": row["sa_q5"],
            "CO6": row["sa_q6"],
        }

        for co_id in co_ids:
            formative = _mean([m1_a, m1_b, m1_t4, m2_a, m2_b, m2_t4])
            theory = _tool_to_scale(theory_scores.get(co_id), theory_max.get(co_id, 10.0))
            summative = _mean([theory, practical])
            final_score = round((formative * 0.5) + (summative * 0.4) + (survey_score * 0.1), 2)
            co_pct = round((final_score / 3.0) * 100.0, 2)

            co_vals[co_id] = co_pct
            co_buckets[co_id]["formative"].append(formative)
            co_buckets[co_id]["summative"].append(summative)
            co_buckets[co_id]["final"].append(final_score)

        overall = round(_mean([co_vals[co_id] for co_id in co_ids]), 2)
        co_vals["Overall_Pct"] = overall
        if overall >= 90:
            grade = "O"
        elif overall >= 80:
            grade = "A+"
        elif overall >= 70:
            grade = "A"
        elif overall >= 60:
            grade = "B+"
        elif overall >= 50:
            grade = "B"
        else:
            grade = "F"
        co_vals["Grade"] = grade
        student_pivot.append(co_vals)

    co_attainment = []
    for co_id in co_ids:
        formative = round(_mean(co_buckets[co_id]["formative"]), 2)
        summative = round(_mean(co_buckets[co_id]["summative"]), 2)
        final = round(_mean(co_buckets[co_id]["final"]), 2)
        pct = round((final / 3.0) * 100.0, 2)
        level = attainment_level(pct)
        final_list = co_buckets[co_id]["final"]
        level1 = sum(1 for score in final_list if score < 1.2)
        level2 = sum(1 for score in final_list if 1.2 <= score < 1.8)
        level3 = sum(1 for score in final_list if score >= 1.8)
        co_attainment.append(
            {
                "CO_ID": co_id,
                "CO_Text": co_text_map.get(co_id, ""),
                "Bloom_Level": bloom_map.get(co_id, ""),
                "IA_Pct": round((formative / 3.0) * 100.0, 2),
                "ETE_Pct": round((summative / 3.0) * 100.0, 2),
                "DA_Score": formative,
                "IDA_Score": survey_score,
                "Final_Score": final,
                "Class_Level": level,
                "Attained": "Attained" if final >= target else "Not Attained",
                "Students_Level1": level1,
                "Students_Level2": level2,
                "Students_Level3": level3,
                "Mean_Pct": pct,
            }
        )

    final_map = {row["CO_ID"]: row["Final_Score"] for row in co_attainment}

    def _build_outcomes(dict_values: Dict[str, str]) -> List[Dict]:
        outcomes = []
        for outcome, description in dict_values.items():
            contributions = []
            for co_id in co_ids:
                strength = matrix_num.get(co_id, {}).get(outcome, 0)
                if strength:
                    contributions.append((co_id, strength))
            if contributions:
                weighted = [final_map[co_id] * (strength / 3.0) for co_id, strength in contributions]
                score = round(sum(weighted) / len(weighted), 2)
            else:
                score = 0.0
            outcomes.append(
                {
                    "outcome": outcome,
                    "description": description,
                    "score": score,
                    "attainment_pct": round((score / 3.0) * 100.0, 2),
                    "level": attainment_level((score / 3.0) * 100.0 if score else 0.0),
                    "contributors": ", ".join(co_id for co_id, _ in contributions),
                }
            )
        return outcomes

    return {
        "report_format": "vfstr_template",
        "vfstr_data": parsed,
        "cos": cos,
        "matrix_label": mapping_doc["matrix_label"],
        "matrix_num": matrix_num,
        "co_attainment": co_attainment,
        "po_attainment": _build_outcomes(po_dict),
        "pso_attainment": _build_outcomes(pso_dict),
        "student_pivot": student_pivot,
        "co_student": [],
        "metadata": {
            "session_id": session["session_id"],
            "course_code": parsed["course_code"] or session.get("course_code", ""),
            "course_name": parsed["course_name"] or session.get("course_name", ""),
            "university": session.get("university", "Vignan's University"),
            "department": session.get("department", ""),
            "degree": session.get("degree", ""),
            "academic_year": session.get("academic_year", ""),
            "semester": session.get("semester", ""),
            "coordinator": parsed["faculty_name"] or session.get("coordinator", ""),
            "num_students": len(parsed["rows"]),
            "target_attainment": target,
            "section": parsed.get("section", ""),
            "survey_score": survey_score,
        },
    }


def generate_vfstr_excel_report(attainment_data: Dict, output_path: str, template_path: str) -> str:
    if not template_path or not os.path.exists(template_path):
        raise FileNotFoundError(f"VFSTR template not found: {template_path}")

    wb = load_workbook(template_path)
    parsed = attainment_data["vfstr_data"]
    meta = attainment_data["metadata"]

    ws_pos = wb["POs"]
    ws_cos = wb["COs"]
    ws_m1 = wb["M1-CO"]
    ws_m2 = wb["M2-CO"]
    ws_sa = wb["SA-CO "]
    ws_survey = wb["Survey"]
    cos = attainment_data.get("cos", [])
    num_cos = max(1, min(6, len(cos)))
    survey_cols = ["E", "F", "G", "H", "I", "J"][:num_cos]

    course_code = parsed["course_code"] or meta.get("course_code", "")
    course_name = parsed["course_name"] or meta.get("course_name", "")
    ws_pos["C5"] = f"{course_code} - {course_name}".strip(" -")
    ws_pos["C6"] = meta.get("semester", "")
    ws_pos["K6"] = parsed.get("section", "") or "1"
    ws_pos["C7"] = meta.get("coordinator", "")
    ws_pos["K7"] = meta.get("academic_year", "")

    for idx, row in enumerate(parsed["rows"], start=33):
        ws_m1[f"C{idx}"] = row["reg_no"]
        ws_m1[f"D{idx}"] = row["m1_qno_a"]
        ws_m1[f"E{idx}"] = row["m1_t1_a"]
        ws_m1[f"O{idx}"] = row["m1_qno_b"]
        ws_m1[f"P{idx}"] = row["m1_t1_b"]
        ws_m1[f"AA{idx}"] = (row["m1_t4"] * 2) if row["m1_t4"] is not None else None

        ws_m2[f"C{idx}"] = row["reg_no"]
        ws_m2[f"D{idx}"] = row["m2_qno_a"]
        ws_m2[f"E{idx}"] = row["m2_t1_a"]
        ws_m2[f"O{idx}"] = row["m2_qno_b"]
        ws_m2[f"P{idx}"] = row["m2_t1_b"]
        ws_m2[f"AA{idx}"] = (row["m2_t4"] * 2) if row["m2_t4"] is not None else None

    sa_map = {
        "D10": 3,
        "E11": 3,
        "F12": 3,
        "G13": 3,
        "H14": 3,
        "I15": 3,
    }
    for cell, value in sa_map.items():
        ws_sa[cell] = value

    for idx, row in enumerate(parsed["rows"], start=29):
        ws_sa[f"C{idx}"] = row["reg_no"]
        ws_sa[f"D{idx}"] = row["sa_q1"]
        ws_sa[f"E{idx}"] = row["sa_q2"]
        ws_sa[f"F{idx}"] = row["sa_q3"]
        ws_sa[f"G{idx}"] = row["sa_q4"]
        ws_sa[f"H{idx}"] = row["sa_q5"]
        ws_sa[f"I{idx}"] = row["sa_q6"]
        ws_sa[f"T{idx}"] = row["sa_practical"]

    survey_value = round(meta.get("survey_score", 0.0), 2)
    ws_survey["D14"] = "Faculty"
    for col in survey_cols:
        ws_survey[f"{col}14"] = survey_value
    for col in ["E", "F", "G", "H", "I", "J"][num_cos:]:
        ws_survey[f"{col}14"] = None

    for idx, row in enumerate(parsed["rows"], start=15):
        ws_survey[f"D{idx}"] = row["reg_no"]
        for col in survey_cols:
            ws_survey[f"{col}{idx}"] = survey_value
        for col in ["E", "F", "G", "H", "I", "J"][num_cos:]:
            ws_survey[f"{col}{idx}"] = None

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(output_path)
    return output_path
